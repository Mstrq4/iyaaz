from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter
from collections.abc import Mapping
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

URL_PATTERN = re.compile(r"(?i)(?:https?://|www\.)\S+")
WHITESPACE_PATTERN = re.compile(r"\s+")
SHEET_NAME = "01_المكتبة_الرئيسية"
SCHEMA_VERSION = 1
SNAPSHOT_NAME = "library.snapshot.json"
MANIFEST_NAME = "library.manifest.json"

FIELD_MAP: tuple[tuple[str, str], ...] = (
    ("الرقم", "id"),
    ("الاختصار", "shortcut"),
    ("الاسم العربي", "nameAr"),
    ("المجال الرئيسي", "mainDomain"),
    ("الفئة", "category"),
    ("الفئة الفرعية", "subcategory"),
    ("نوع الاختصار", "shortcutType"),
    ("الوظيفة", "functionText"),
    ("المدخلات المطلوبة", "requiredInputs"),
    ("تعليمات التنفيذ المختصرة", "executionInstructions"),
    ("المخرجات", "outputs"),
    ("المقاس / النسبة", "sizeRatio"),
    ("الخامات / التقنية", "materialsTech"),
    ("الإضاءة", "lighting"),
    ("التثبيت / التنفيذ", "installationExecution"),
    ("الأسلوب البصري", "visualStyle"),
    ("قاعدة الالتزام بالهوية", "brandCompliance"),
    ("الاختصارات المدمجة", "combinedShortcuts"),
    ("أفضل استخدام", "bestUse"),
    ("كلمات مفتاحية", "keywords"),
    ("نوع الأصل", "assetType"),
    ("ملاحظات", "notes"),
)

TEXT_KEYS = tuple(target for _, target in FIELD_MAP if target != "id")
PUBLIC_KEYS = tuple(target for _, target in FIELD_MAP)


def sanitize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = URL_PATTERN.sub("", text)
    return WHITESPACE_PATTERN.sub(" ", text).strip()


def normalize_id(value: Any) -> int:
    if isinstance(value, bool):
        raise ValueError("id must be a positive integer")
    if isinstance(value, int):
        number = value
    elif isinstance(value, float) and value.is_integer():
        number = int(value)
    else:
        text = str(value).strip()
        if not text.isdigit():
            raise ValueError("id must be a positive integer")
        number = int(text)
    if number <= 0:
        raise ValueError("id must be a positive integer")
    return number


def sanitize_row(row: Mapping[str, Any]) -> dict[str, Any]:
    record: dict[str, Any] = {}
    for source_key, target_key in FIELD_MAP:
        raw_value = row.get(source_key)
        record[target_key] = normalize_id(raw_value) if target_key == "id" else sanitize_text(raw_value)
    if not record["shortcut"]:
        raise ValueError("shortcut is required")
    if not record["nameAr"]:
        raise ValueError("nameAr is required")
    assert_safe_record(record)
    return record


def serialize_record(record: Mapping[str, Any]) -> str:
    return json.dumps(dict(record), ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def assert_safe_record(record: Mapping[str, Any]) -> None:
    unexpected = set(record) - set(PUBLIC_KEYS)
    if unexpected:
        raise ValueError(f"unexpected public fields: {sorted(unexpected)}")
    if URL_PATTERN.search(serialize_record(record)):
        raise ValueError("public record contains a URL")


def load_public_records(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"missing worksheet: {SHEET_NAME}")
    worksheet = workbook[SHEET_NAME]
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    except StopIteration as exc:
        raise ValueError("library worksheet is empty") from exc
    missing = [source for source, _ in FIELD_MAP if source not in headers]
    if missing:
        raise ValueError(f"missing required columns: {missing}")

    records: list[dict[str, Any]] = []
    for values in rows:
        if not any(value not in (None, "") for value in values):
            continue
        row = dict(zip(headers, values, strict=False))
        records.append(sanitize_row(row))
    validate_snapshot(records)
    return records


def validate_snapshot(records: list[dict[str, Any]]) -> None:
    if len(records) != 5812:
        raise ValueError(f"expected 5812 records, got {len(records)}")
    ids = [record["id"] for record in records]
    if ids != list(range(1, 5813)):
        raise ValueError("record ids must be contiguous from 1 through 5812")
    shortcuts = [record["shortcut"] for record in records]
    if len(set(shortcuts)) != len(shortcuts):
        raise ValueError("shortcuts must be unique")
    for record in records:
        assert_safe_record(record)


def encode_records(records: list[dict[str, Any]]) -> bytes:
    return json.dumps(records, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")


def write_snapshot(workbook_path: Path, output_dir: Path) -> dict[str, Any]:
    records = load_public_records(workbook_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    raw = encode_records(records)
    (output_dir / SNAPSHOT_NAME).write_bytes(raw)
    type_counts = Counter(str(record["shortcutType"]) for record in records)
    manifest = {
        "schemaVersion": SCHEMA_VERSION,
        "format": "json",
        "file": SNAPSHOT_NAME,
        "recordCount": len(records),
        "domainCount": len({record["mainDomain"] for record in records}),
        "categoryCount": len({record["category"] for record in records}),
        "subcategoryCount": len({record["subcategory"] for record in records}),
        "typeCounts": dict(sorted(type_counts.items())),
        "publicFields": list(PUBLIC_KEYS),
        "sha256": hashlib.sha256(raw).hexdigest(),
    }
    (output_dir / MANIFEST_NAME).write_text(
        json.dumps(manifest, ensure_ascii=False, sort_keys=True, indent=2) + "\n",
        encoding="utf-8",
    )
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the sanitized IYAAZ runtime JSON snapshot.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output-dir", type=Path, default=Path("data"))
    args = parser.parse_args()
    manifest = write_snapshot(args.workbook, args.output_dir)
    print(json.dumps(manifest, ensure_ascii=False, sort_keys=True))


if __name__ == "__main__":
    main()


__all__ = [
    "FIELD_MAP",
    "MANIFEST_NAME",
    "PUBLIC_KEYS",
    "SCHEMA_VERSION",
    "SHEET_NAME",
    "SNAPSHOT_NAME",
    "TEXT_KEYS",
    "assert_safe_record",
    "encode_records",
    "load_public_records",
    "normalize_id",
    "sanitize_row",
    "sanitize_text",
    "serialize_record",
    "validate_snapshot",
    "write_snapshot",
]
